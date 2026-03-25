# coding = utf8
import ctypes
import gc
import json
import logging
import os
import random
import re
import subprocess
import sys
import threading
import time
import xml.etree.cElementTree as et
from time import sleep

import imagehash
import openpyxl
import pandas as pd
import serial
from PIL import Image
from airtest.core.api import connect_device
from poco.drivers.android.uiautomation import AndroidUiautomationPoco
from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from serial.tools.list_ports_windows import comports
from webdriver_manager.microsoft import EdgeChromiumDriverManager

logger_airtest = logging.getLogger("airtest")
logger_airtest.setLevel(logging.ERROR)
os.path.abspath("../SKDL0104")

"""
    @Project:pythonProject_seevision
    @File:common_api_SMD0302.py
    @Author:十二点前要睡觉
    @Date:2023/3/6 18:06
"""

"""
    Android Device Common Api
"""
# 继电器串口1~8个开关的串口信息
RELAY_CONTROL_COMPORT_1_OPEN = [0xFE, 0x05, 0x00, 0x00, 0xFF, 0x00, 0x98, 0x35]
RELAY_CONTROL_COMPORT_1_CLOSE = [0xFE, 0x05, 0x00, 0x00, 0x00, 0x00, 0xD9, 0xC5]

RELAY_CONTROL_COMPORT_2_OPEN = [0xFE, 0x05, 0x00, 0x01, 0xFF, 0x00, 0xC9, 0xF5]
RELAY_CONTROL_COMPORT_2_CLOSE = [0xFE, 0x05, 0x00, 0x01, 0x00, 0x00, 0x88, 0x05]

RELAY_CONTROL_COMPORT_3_OPEN = [0xFE, 0x05, 0x00, 0x02, 0xFF, 0x00, 0x39, 0xF5]
RELAY_CONTROL_COMPORT_3_CLOSE = [0xFE, 0x05, 0x00, 0x02, 0x00, 0x00, 0x78, 0x05]

RELAY_CONTROL_COMPORT_4_OPEN = [0xFE, 0x05, 0x00, 0x03, 0xFF, 0x00, 0x68, 0x35]
RELAY_CONTROL_COMPORT_4_CLOSE = [0xFE, 0x05, 0x00, 0x03, 0x00, 0x00, 0x29, 0xC5]

RELAY_CONTROL_COMPORT_5_OPEN = [0xFE, 0x05, 0x00, 0x04, 0xFF, 0x00, 0xD9, 0xF4]
RELAY_CONTROL_COMPORT_5_CLOSE = [0xFE, 0x05, 0x00, 0x04, 0x00, 0x00, 0x98, 0x04]

RELAY_CONTROL_COMPORT_6_OPEN = [0xFE, 0x05, 0x00, 0x05, 0xFF, 0x00, 0x88, 0x34]
RELAY_CONTROL_COMPORT_6_CLOSE = [0xFE, 0x05, 0x00, 0x05, 0x00, 0x00, 0xC9, 0xC4]

RELAY_CONTROL_COMPORT_7_OPEN = [0xFE, 0x05, 0x00, 0x06, 0xFF, 0x00, 0x78, 0x34]
RELAY_CONTROL_COMPORT_7_CLOSE = [0xFE, 0x05, 0x00, 0x06, 0x00, 0x00, 0x39, 0xC4]

RELAY_CONTROL_COMPORT_8_OPEN = [0xFE, 0x05, 0x00, 0x07, 0xFF, 0x00, 0x29, 0xF4]
RELAY_CONTROL_COMPORT_8_CLOSE = [0xFE, 0x05, 0x00, 0x07, 0x00, 0x00, 0x68, 0x04]

RELAY_CONTROL_COMPORT_9_OPEN = [0xFE, 0x05, 0x00, 0x08, 0xFF, 0x00, 0x19, 0xF7]
RELAY_CONTROL_COMPORT_9_CLOSE = [0xFE, 0x05, 0x00, 0x08, 0x00, 0x00, 0x58, 0x07]

RELAY_CONTROL_COMPORT_10_OPEN = [0xFE, 0x05, 0x00, 0x09, 0xFF, 0x00, 0x48, 0x37]
RELAY_CONTROL_COMPORT_10_CLOSE = [0xFE, 0x05, 0x00, 0x09, 0x00, 0x00, 0x09, 0xC7]


# 通用安卓设备的API
class Common_Api:

    def __init__(self, deviceSN, log_name):
        self.deviceSN = deviceSN
        self.log_name = log_name
        self.device, self.poco = self.initAirtestDevicePoco()

    # 执行cmd指令
    def runCMDCommand(self, command):
        popen = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)
        backData = ""
        try:
            if "monkey" in command or "fastboot" in command or "logcat" in command or "install" in command or "push" in command:
                temp_data = popen.communicate()[0]
            else:
                temp_data = popen.communicate(timeout=3)[0]
            if temp_data:
                backData = temp_data.decode("utf-8", errors="ignore")
        # except (Exception or PermissionError):
        except Exception as ex:
            # popen.kill()
            # print(str(ex))
            print("Warning ~")
        # backData = str(
        #     subprocess.Popen(command, shell=True, stdout=subprocess.PIPE).communicate()[0].decode("utf-8"))
        if backData:
            # print(backData)
            self.toTxt(backData)
        # print(backData)
        # self.toTxt(backData)
        return backData

    # 日志写入到指定txt文件
    def toTxt(self, result):
        try:
            print(result)
            with open("{}/[{}]{}SerialPortTestLog.log".format(os.path.dirname(os.path.realpath(__file__)),
                                                              self.deviceSN, self.log_name), "a+") as f:
                cur_time = time.strftime("%Y%m%d_%H%M%S")
                result = result.encode("utf-8", errors="ignore").decode("utf-8", errors="ignore")
                f.write("【{}】-【device - {}】 - ".format(cur_time, self.deviceSN) + result + "\n")
        except (AttributeError, TypeError) as ex:
            print("【Device{}】 - ".format(self.deviceSN) + "【Error need check, maybe not important】 : \r\n{}\r\n".format(
                str(ex)))
            f.write(
                "【Device{}】 - ".format(self.deviceSN) + "【Error need check, maybe not important】 : \r\n{}\r\n".format(
                    str(ex)))

    # 获取设备当前版本
    def getCurDeviceVersion(self):
        version = self.runCMDCommand('adb -s {} shell "getprop | grep version"'.format(self.deviceSN))
        regex = r"\[ro.bootimage.build.version.incremental\]: \[(.*)\]"
        curVersion = str(re.findall(regex, version)[0]).strip()
        return curVersion

    # 启动线程运行logcat
    def logcat(self):
        self.runCMDCommand(
            "adb -s {} logcat -b all > [{}]_{}_logcat.log".format(self.deviceSN, self.deviceSN, self.log_name))

    # 启动logcat线程后台运行
    def runLogcatThread(self):
        t_logcat = multi_thread(target=self.logcat)
        t_logcat.start()
        return t_logcat

    # 列出app的未授权权限,筛选掉无需授权的应用权限
    def list_permission(self, package_name):
        # 判断没有启动界面就放出来进行授权
        # 优化后app数量:67个，优化前app数量:330个
        global permission_list

        if str(self.runCMDCommand(
                'adb -s {} shell "dumpsys package {} | grep category.LAUNCHER"'.format(self.deviceSN,
                                                                                       package_name))).replace(" ",
                                                                                                               "").replace(
            " ", "").replace(
            "b''",
            "") is not None:
            permission_list = str(
                self.runCMDCommand(
                    'adb -s {} shell "dumpsys package {} | grep permission | grep granted=false"'.format(self.deviceSN,
                                                                                                         package_name))
            )
            permission_list = re.findall("\s*([a-zA-Z0-9_.]*):\sgranted", permission_list)
            if len(permission_list) == 0:
                print("Current app is no need to authorized！")
            else:
                pass
        return permission_list

    # 列出有launcher界面的app
    def controlAppRange(self):
        wholeApp = self.runCMDCommand("adb -s {} shell pm list packages".format(self.deviceSN))
        app_list = re.findall("package:(.*)", str(wholeApp).replace("\\r\\n", "\n").replace("\r", ""))
        launchableApp_list = []
        for package in app_list:
            haveLauncher = str(
                self.runCMDCommand(
                    'adb -s {} shell "dumpsys package {} | grep category.LAUNCHER"'.format(self.deviceSN, package))
            ).replace(" ", "").replace("b''", "")
            if haveLauncher:
                # print("APP 【{}】 have launcher page to view!".format(package))
                launchableApp_list.append(package)
        print("Before filter，Current system has app:{}".format(len(app_list)))
        print("After filter，Current system has app:{}".format(len(launchableApp_list)))
        print("Test app list contains:\n{}".format(launchableApp_list))
        return launchableApp_list

    # 进行授权操作
    def grant_permission(self, app_list):
        print("Device is authorizing, please waiting for a moment……")
        app_permission = self.data_deal(app_list)
        for app_ in app_permission:
            for permission_ in app_[1]:
                try:
                    print("Now app:【{}】 - 【{}】 permission authorized done！".format(app_[0], permission_))
                    os.system("adb -s {} shell pm grant {} {}".format(self.deviceSN, app_[0], permission_))
                except Exception:
                    print("This permission cannot be authorize, skip it！")

    # 列出app需要授权的所有权限
    def data_deal(self, app_list):
        app_permission = []
        for package_name in app_list:
            permission_name = self.list_permission(package_name)
            app_permission.append([package_name, permission_name])
        return app_permission

    # 清除app的缓存
    def clearAppDataBuffer(self, packageName="com.tencent.wemeet.app"):
        self.runCMDCommand("adb -s {} shell pm clear {}".format(self.deviceSN, packageName))
        print("APP {} data clear finished".format(packageName))

    # 获取当前实时时间戳 - 毫秒级
    def getMSTime(self):
        t = time.time()
        return int(round(t * 1000))

    # 通过shell input tap去点击打开App
    def touchApp(self, x_pos, y_pos, activityName="com.tencent.wemeet.app.StartupActivity"):
        time1 = self.getMSTime()
        print(time1)
        self.runCMDCommand("adb -s {} shell input tap {} {}".format(self.deviceSN, x_pos, y_pos))
        return self.keepWaitingActivity(time1=time1, activityName=activityName)

    # dump下当前设备界面的ui tree
    def dumpXml(self):
        self.runCMDCommand(
            f'adb -s {"{}"} shell uiautomator dump --compressed /{"/sdcard/ui.xml"}'.format(self.deviceSN))
        self.runCMDCommand(f'adb -s {"{}"} pull {"/sdcard/ui.xml"} {"./"}'.format(self.deviceSN))
        source = et.parse("./ui.xml")
        return source.getroot()

    def getCurrentPageApps(self):
        applist = []
        for app in self.poco("com.youdao.hardware.panda:id/vpMain").offspring(
                name="com.youdao.hardware.panda:id/appName"):
            applist.append(app.get_text())
        for foloer in self.poco("com.youdao.hardware.panda:id/vpMain").offspring(
                name="com.youdao.hardware.panda:id/foldName"):
            applist.append(foloer.get_text())
        return applist

    def dragAppIconTo(self, app_label):
        try:
            self.poco(text=app_label).start_gesture().hold(0.5).to(
                (random.uniform(0.1, 0.8), random.uniform(0.1, 0.8))).hold(0.5).up()
            self.poco(text="完成").wait(3).click()
            return True
        except Exception as ex:
            self.toTxt("发生错误，停止测试：\n{}".format(str(ex)))
            return False

    # 获取App icon的中心点坐标（用于点击）
    def getAppCenteralPosition(self, app_text):
        root = self.dumpXml()
        for node in root.iter("node"):
            if node.attrib["text"] == app_text:
                bounds = node.attrib["bounds"]
                pattern = re.compile(r"\d+")
                coord = pattern.findall(bounds)
                x_pos = (int(coord[2]) - int(coord[0])) / 2.0 + int(coord[0])
                y_pos = (int(coord[3]) - int(coord[1])) / 2.0 + int(coord[1])
                return x_pos, y_pos

    # 等待App界面启动完成
    def keepWaitingActivity(self, activityName="com.tencent.wemeet.sdk.meeting.premeeting.home.GuestGuideActivity",
                            time1=0):
        while True:
            returnLine = str(
                self.runCMDCommand('adb -s {} shell "dumpsys window | grep mCurrentFocus"'.format(self.deviceSN))
            ).replace("b'", "").replace("\\n'", "")
            print(returnLine)
            # 特殊情况3:微信多次kill会触发安全模式，需要兼容下再Kill掉
            if "com.tencent.mm.recovery.ui.RecoveryUI" in returnLine:
                self.killApp("com.tencent.mm")
                return 0
            if activityName in returnLine:
                print("APP boot finished")
                time2 = self.getMSTime()
                print("time1 = {} and time2 = {}".format(time1, time2))
                return time2 - time1

    # 杀掉某个app
    def killApp(self, package):
        command = "adb -s {} shell am force-stop {}".format(self.deviceSN, package)
        self.runCMDCommand(command)

    # 初始化安卓设备Airtest device和poco
    def initAirtestDevicePoco(self):
        device = connect_device("Android:///{}".format(self.deviceSN))
        device.stop_app("com.netease.open.pocoservice")
        poco = AndroidUiautomationPoco(device=device, use_airtest_input=False, screenshot_each_action=False)
        return device, poco

    # adb shell - home
    def adb_home(self):
        self.runCMDCommand("adb -s {} shell input keyevent 3".format(self.deviceSN))

    def adb_back(self):
        self.runCMDCommand("adb -s {} shell input keyevent 4".format(self.deviceSN))

    # adb shell - get bluetooth status
    def getBluetoothStatus(self):
        bluetoothStatus = self.runCMDCommand("adb -s {} shell settings get global bluetooth_on".format(self.deviceSN))
        # print(bluetoothStatus)
        return str(bluetoothStatus).replace("b'", "").replace("\\r\\n'", "").strip()

    # adb shell - enable bluetooth
    def bluetoothOn(self):
        var = self.runCMDCommand("adb -s {} shell svc bluetooth enable".format(self.deviceSN))
        sleep(1)
        if self.getBluetoothStatus() == "1":
            return "success"
        else:
            return "fail"

    # adb shell - disable bluetooth
    def bluetoothOff(self):
        var = self.runCMDCommand("adb -s {} shell svc bluetooth disable".format(self.deviceSN))
        sleep(1)
        if self.getBluetoothStatus() == "0":
            return "success"
        else:
            return "fail"

    # adb shell screencap - catch device current page
    def screenshotAndPullOut(self, folder_name, picName):
        if not os.path.exists("./{}/".format(folder_name)):
            os.mkdir("./{}/".format(folder_name))
        print("adb -s {} screencap begin, current page will be catched……".format(self.deviceSN))
        cur_time = time.strftime("%Y%m%d_%H%M%S")
        self.runCMDCommand("adb -s {} shell screencap -p /sdcard/[{}]{}.png".format(self.deviceSN, cur_time, picName))
        sleep(3)
        print("screencap is done, picture {} is pulled out.".format(picName))
        self.runCMDCommand(
            "adb -s {} pull /sdcard/[{}]{}.png ./{}/".format(self.deviceSN, cur_time, picName, folder_name))
        sleep(3)
        return "[{}]{}.png".format(cur_time, picName)

    # adb shell - get wifi status

    def getWifiStatus(self):
        wifiStatus = self.runCMDCommand("adb -s {} shell settings get global wifi_on".format(self.deviceSN))
        # print(wifiStatus)
        return str(wifiStatus).replace("b'", "").replace("\\r\\n'", "").strip()

    # adb shell - enable wifi
    def wifiOn(self):
        var = self.runCMDCommand("adb -s {} shell svc wifi enable".format(self.deviceSN))
        sleep(1)
        if self.getWifiStatus() == "2":
            return "success"
        else:
            return "fail"

    # adb shell - disable wifi
    def wifiOff(self):
        var = self.runCMDCommand("adb -s {} shell svc wifi disable".format(self.deviceSN))
        sleep(1)
        if self.getWifiStatus() == "0":
            return "success"
        else:
            return "fail"

    def picture_compare(self, original_image, compare_image):
        original = imagehash.average_hash(Image.open(original_image))
        compare = imagehash.average_hash(Image.open(compare_image))
        gap_value = compare - original
        self.toTxt("gap_value is [{}], less than 100 will pass".format(gap_value))
        if gap_value <= 100:
            result = "{} compare Test result is [PASS]".format(compare_image)
        else:
            result = "{} compare Test result is [FAIL]".format(compare_image)
        return result

    def picture_compare(self, original_image, compare_image, picture_compare_standrad_value):
        original = imagehash.average_hash(Image.open(original_image))
        compare = imagehash.average_hash(Image.open(compare_image))
        gap_value = compare - original
        self.toTxt("gap_value is [{}], less than {} will pass".format(gap_value, picture_compare_standrad_value))
        if gap_value <= picture_compare_standrad_value:
            result = "{} compare Test result is [PASS]".format(compare_image)
        else:
            result = "{} compare Test result is [FAIL]".format(compare_image)
        return result

    # set no screen off
    def setNoSleep(self):
        self.runCMDCommand(
            "adb -s {} shell settings put system screen_off_timeout 2147483647 ".format(self.deviceSN))

    # unlock screen
    def unlockScreen(self):
        self.setNoSleep()
        self.device.wake()
        self.device.unlock()
        self.device.home()

    # adb install -g xxx.apk
    def installAllApk_GrantPermission(self):
        os.chdir(os.path.dirname(os.path.realpath(__file__)))
        apks = os.listdir("./apk/")
        for apk in apks:
            self.toTxt("正在安裝{}".format(apk))
            self.runCMDCommand("adb -s {} install -g ./apk/{}".format(self.deviceSN, apk))
        self.toTxt("安裝完成")

    # slide 滑动屏幕
    def slideScreen(self, direction="left"):
        """
        左滑：从右到左滑动屏幕，根据屏幕宽高swipe或百分比scroll进行适配
        :param direction:
        :return:
        """
        if direction == "left":
            self.poco.scroll(direction="horizontal", percent=0.4, duration=0.06)
        elif direction == "right":
            self.poco.scroll(direction="horizontal", percent=-0.4, duration=0.06)
        elif direction == "up":
            self.poco.scroll(direction="vertical", percent=0.6, duration=0.6)
        elif direction == "down":
            self.poco.scroll(direction="vertical", percent=-0.6, duration=0.6)

    def dumpCurrentScreen(self):
        return self.runCMDCommand("adb -s {} shell dumpsys window | grep mCurrentFocus".format(self.deviceSN))

    def setOrientation(self, mode):
        self.toTxt("Set Orientation to {}!".format(mode))
        self.runCMDCommand(
            "adb -s {} shell content insert --uri content://settings/system --bind name:s:accelerometer_rotation --bind value:i:0".format(
                self.deviceSN))
        if mode == "horizontal":
            self.runCMDCommand(
                "adb -s {} shell content insert --uri content://settings/system --bind name:s:user_rotation --bind value:i:1".format(
                    self.deviceSN))
        elif mode == "vertical":
            self.runCMDCommand(
                "adb -s {} shell content insert --uri content://settings/system --bind name:s:user_rotation --bind value:i:0".format(
                    self.deviceSN))

    def checkOnLauncher(self):
        if "com.youdao.hardware.panda" in self.runCMDCommand(
                "adb -s {} shell dumpsys window | grep mCurrentFocus".format(self.deviceSN)):
            return True
        else:
            return False

    """
           @description:滚动查找元素
           @param:
               element_text:元素text属性
               element_id:元素id属性
        """

    def scroll_to_find_element(self, element_text="", element_id=""):
        self.toTxt("function:" + sys._getframe().f_code.co_name + ":滚动查找元素:")
        global element
        menu_exists = False
        search_count = 0
        if element_text != "":
            while not menu_exists:
                element = self.poco(text=element_text).wait()
                menu_exists = element.exists()
                if menu_exists:
                    return element
                self.poco.scroll(direction="vertical", percent=0.6, duration=1)
                search_count += 1
                # 给滑动查找增加向上滑动，兼容到底未找到的情况，即向下查找超过10次则开始向上查找
                while search_count >= 5 and not menu_exists:
                    self.poco.scroll(direction="vertical", percent=-0.6, duration=1)
                    element = self.poco(text=element_text).wait()
                    menu_exists = element.exists()
                    search_count += 1
                    if search_count >= 10:
                        search_count = 0
                        break
                    if menu_exists:
                        return element
        else:
            while not menu_exists:
                element = self.poco(element_id).wait()
                menu_exists = element.exists()
                if menu_exists:
                    return element
                self.poco.scroll(direction="vertical", percent=0.6, duration=1)
                search_count += 1
                # 给滑动查找增加向上滑动，兼容到底未找到的情况，即向下查找超过10次则开始向上查找
                while search_count >= 5 and not menu_exists:
                    self.poco.scroll(direction="vertical", percent=-0.6, duration=1)
                    element = self.poco(element_id).wait()
                    menu_exists = element.exists()
                    search_count += 1
                    if search_count >= 10:
                        search_count = 0
                        break
                    if menu_exists:
                        return element
        return element


class selenium_single_object:
    def __init__(self, enter_url, log_name):
        """
        初始化Selenium自动化测试对象
        :param enter_url: 第一次需要打开的链接
        :param log_name: 测试脚本名称
        """
        self.enter_url = enter_url
        self.log_name = log_name
        self.log_saveName = str(self.enter_url).replace(":", "_").replace("/", "_").replace(".", "_")
        self.driver = self.initWebDriver()

    def initWebDriver(self):
        """
        初始化WebDriver，有界面操作，Edge浏览器 (离线模式)
        """
        self.toTxt("初始化浏览器自动化对象，即将访问设备url：{}".format(self.enter_url))

        try:
            option = webdriver.EdgeOptions()
            option.add_experimental_option("detach", True)
            option.add_argument('--start-maximized')

            # === 修改部分开始：使用本地驱动 ===

            # 获取当前脚本所在目录的 msedgedriver.exe 路径
            current_dir = os.path.dirname(os.path.abspath(__file__))
            driver_path = os.path.join(current_dir, "msedgedriver.exe")

            self.toTxt(f"正在尝试加载本地驱动: {driver_path}")

            if not os.path.exists(driver_path):
                raise FileNotFoundError(f"找不到驱动文件，请确保 msedgedriver.exe 在 {current_dir} 目录下")

            service = Service(executable_path=driver_path)
            driver = webdriver.Edge(options=option, service=service)

            # === 修改部分结束 ===

            driver.implicitly_wait(3)
            self.toTxt("浏览器初始化完成，当前是【有界面】自动化操作……")
            return driver

        except Exception as e:
            self.toTxt(f"启动浏览器失败！错误信息: {e}")
            raise e

    def openUrl(self, url_enter):
        """
        进入指定链接
        :param url_enter: 进入的链接
        :return:
        """
        self.toTxt("进入{}链接……".format(url_enter))
        self.driver.get(url_enter)
        self.driver.maximize_window()
        waitTimeBack(1)
        self.toTxt("进入成功！")

    # ------------------------ SXW0301Project Control Start ------------------------
    def loginGBS(self):
        """
        登录GBS服务器
        :return:
        """
        self.toTxt("登录GBS服务器")
        useraccount = self.driver.find_element(By.NAME, "user")
        password = self.driver.find_element(By.NAME, "pwd")
        loginBtn = self.driver.find_element(By.CLASS_NAME, "sign")
        useraccount.send_keys("sxtest")
        password.send_keys("9db70952")
        loginBtn.click()
        waitTimeBack(1)
        self.toTxt("登录GBS服务器成功！")

    def loginDeviceSettingsPage(self):
        """
        登录设备的内网配置页面
        :return:
        """
        self.toTxt("登录设备内网设置页面")
        try:
            accountBtn = self.driver.find_element(By.CLASS_NAME, "el-avatar--circle")
            accountBtn.click()
            waitTimeBack(1)
        except Exception:
            self.toTxt("无需点击头像登录！")
        useraccount = self.driver.find_element(By.XPATH,
                                               '//input[@type="text" and contains(@class, "el-input__inner")]')
        password = self.driver.find_element(By.XPATH,
                                            '//input[@type="password" and contains(@class, "el-input__inner")]')
        useraccount.send_keys("admin")
        password.send_keys("123456")
        loginBtn = self.driver.find_element(By.CLASS_NAME, "login-btn")
        loginBtn.click()
        wait = WebDriverWait(self.driver, 10)
        alert = wait.until(EC.alert_is_present())
        alert.accept()
        waitTimeBack(1)
        self.toTxt("设备内网设备页面登录完成！")

    def getDeviceSettingsIPCID(self):
        """
        从设备内网配置页面获取设备的IPCID
        :return:返回ipcid值
        """
        self.toTxt("获取当前设备的IPCID")
        self.driver.find_element(By.XPATH, "//span[text()='网络配置']").click()
        self.scrollToSpecificElement("xpath", "//span[text()='IPC设备ID']")
        ipcIDInput = self.driver.find_element(By.XPATH,
                                              '//span[text()="IPC设备ID"]/following-sibling::div/input['
                                              '@class="el-input__inner"]')
        waitTimeBack(1)
        ipcid = ipcIDInput.get_attribute("value")
        self.toTxt("当前设备IPCID为：{}".format(ipcid))
        return ipcid

    def rebootDeviceBySettingsPage(self):
        self.toTxt("通过设备设置页面对设备进行重启……")
        self.driver.find_element(By.XPATH, "//span[text()='系统配置']").click()
        rebootBtn = self.driver.find_element(By.XPATH,
                                             '//button[@type="button" and contains(@class, "el-button butt '
                                             'el-button--default")]/span[text()="重启系统"]')
        rebootBtn.click()
        waitTimeBack(1)
        confirmBtn = self.driver.find_element(By.XPATH,
                                              '//button[@type="button" and contains(@class, "el-button '
                                              'el-button--default el-button--small el-button--primary ")]/span['
                                              'contains(text(), "确定")]')
        confirmBtn.click()

    def resetDeviceBySettingsPage(self):
        self.toTxt("通过设备设置页面对设备进行重置……")
        self.driver.find_element(By.XPATH, "//span[text()='系统配置']").click()
        resetBtn = self.driver.find_element(By.XPATH,
                                            '//button[@type="button" and contains(@class, "el-button butt '
                                            'el-button--default")]/span[text()="恢复出厂设置"]')
        resetBtn.click()
        waitTimeBack(1)
        confirmBtn = self.driver.find_element(By.XPATH,
                                              '//button[@type="button" and contains(@class, "el-button '
                                              'el-button--default el-button--small el-button--primary ")]/span['
                                              'contains(text(), "确定")]')
        confirmBtn.click()

    def enableGB28181FuncBySettingsPage(self):
        self.toTxt("通过设备设置页面 - 打开GB28181功能")
        self.driver.find_element(By.XPATH, "//span[text()='网络配置']").click()
        self.scrollToSpecificElement("xpath", "//span[text()='IPC设备ID']")
        waitTimeBack(1)
        enable28181RadioBtn = self.driver.find_elements(By.XPATH, "//span[contains(text(), '开启')]")[3]
        enable28181RadioBtn.click()
        confirm28181Btn = self.driver.find_elements(By.XPATH, "//span[text()='应用']")[3]
        confirm28181Btn.click()
        waitTimeBack(3)
        wait = WebDriverWait(self.driver, 10)
        alert = wait.until(EC.alert_is_present())
        alert.accept()
        waitTimeBack(1)
        self.toTxt("GB28181功能打开成功！")

    def disableGB28181FuncBySettingsPage(self):
        self.toTxt("通过设备设置页面 - 关闭GB28181功能")
        self.driver.find_element(By.XPATH, "//span[text()='网络配置']").click()
        self.scrollToSpecificElement("xpath", "//span[text()='IPC设备ID']")
        waitTimeBack(1)
        disable28181RadioBtn = self.driver.find_elements(By.XPATH, "//span[contains(text(), '关闭')]")[3]
        disable28181RadioBtn.click()
        confirm28181Btn = self.driver.find_elements(By.XPATH, "//span[text()='应用']")[3]
        confirm28181Btn.click()
        waitTimeBack(3)
        wait = WebDriverWait(self.driver, 10)
        alert = wait.until(EC.alert_is_present())
        alert.accept()
        waitTimeBack(1)
        self.toTxt("GB28181功能关闭成功！")

    def enterDeviceGBSVideoStreamControlPage(self, ipcid):
        """
        进入设备GBS的视频流控制页面
        :param ipcid:
        :return:
        """
        self.toTxt("进入设备GBS视频流控制页面")
        productionLink = self.driver.find_element(By.XPATH, "//a[contains(text(), '产品')]")
        productionLink.click()
        inTimeAudioVideoLink = self.driver.find_element(By.XPATH, "//a[contains(text(), '实时音视频')]")
        inTimeAudioVideoLink.click()
        waitTimeBack(3)
        self.driver.switch_to.frame("cloud")
        waitTimeBack(2)
        GBS_CheckBtn = self.driver.find_element(By.XPATH,
                                                '//td[contains(text(), "课程质量分析('
                                                'GBS)")]/following-sibling::td/button/a[contains(text(), "查看")]')
        GBS_CheckBtn.click()
        waitTimeBack(1)
        operationManageLink = self.driver.find_element(By.XPATH, "//a[contains(text(), '运营管理')]")
        operationManageLink.click()
        self.driver.switch_to.default_content()
        waitTimeBack(1)
        self.driver.switch_to.frame("cloud")
        self.driver.switch_to.frame("rtc")
        self.driver.switch_to.frame("right")
        self.toTxt("切到iframe成功")
        ipcidInput = self.driver.find_element(By.NAME, 'deviceId')
        ipcidInput.send_keys(ipcid)
        findBtn = self.driver.find_element(By.CLASS_NAME, "btn-primary")
        findBtn.click()
        waitTimeBack(1)
        if self.checkDeviceStatusGBS():
            GBS_CheckBtn_2 = self.driver.find_element(By.XPATH,
                                                      '//td[contains(text(), "{}")]/following-sibling::td/button/a[contains(text(), "查看")]'.format(
                                                          ipcid))
            GBS_CheckBtn_2.click()
            return True
        else:
            self.toTxt("停止测试，设备已经离线，请检查设备状态！")
            return False

    def openChannel_1_window(self):
        """
        打开GBS设备Channel 1的播放窗口
        :return:
        """
        self.toTxt("打开Channel 1窗口")
        Channel1Btn = self.driver.find_elements(By.XPATH,
                                                "//td[text()='ON']/following-sibling::td/button[contains(text(), '播放')]")[
            0]
        Channel1Btn.click()
        waitTimeBack(1)
        self.toTxt("Channel 1播放窗口打开完成！")

    def openChannel_2_window(self):
        """
        打开GBS设备Channel 2的播放窗口
        :return:
        """
        self.toTxt("打开Channel 2窗口")
        Channel2Btn = self.driver.find_elements(By.XPATH,
                                                "//td[text()='ON']/following-sibling::td/button[contains(text(), '播放')]")[
            1]
        Channel2Btn.click()
        waitTimeBack(1)
        self.toTxt("Channel 2播放窗口打开完成！")

    def Channel_1_Play(self):
        """
        GBS设备Channel 1进行播放
        :return: 返回播放结果
        """
        self.toTxt("Channel 1 视频流开始播放！")
        channel_1_playBtn = self.driver.find_element(By.ID, "flv")
        channel_1_playBtn.click()
        waitTimeBack(10)
        # 定位到 <video> 元素
        video_element = self.driver.find_element(By.TAG_NAME, 'video')
        # 使用 JavaScript 获取当前播放时间
        current_time = self.driver.execute_script('return arguments[0].currentTime;', video_element)
        # 打印当前播放时间
        self.toTxt(f'当前播放时间：{current_time} 秒')
        if str(current_time) != "0":
            self.toTxt("播放完成！")
            return True
        else:
            self.toTxt("播放失败！")
            return False

    def Channel_2_Play(self):
        """
        GBS设备Channel 2进行播放
        :return: 返回播放结果
        """
        self.toTxt("Channel 2 视频流开始播放！")
        channel_2_playBtn = self.driver.find_element(By.ID, "flv")
        channel_2_playBtn.click()
        waitTimeBack(20)
        # 定位到 <video> 元素
        video_element = self.driver.find_element(By.TAG_NAME, 'video')
        # 使用 JavaScript 获取当前播放时间
        current_time = self.driver.execute_script('return arguments[0].currentTime;', video_element)
        # 打印当前播放时间
        self.toTxt(f'当前播放时间：{current_time} 秒')
        if str(current_time) != "0":
            self.toTxt("播放完成！")
            return True
        else:
            self.toTxt("播放失败！")
            return False

    def gbs_refresh(self):
        self.toTxt("GBS拉流页面刷新")
        refreshBtn = self.driver.find_element(By.XPATH, "//button[contains(text(), '刷新')]")
        refreshBtn.click()

    def VideoChannel_1_Close(self):
        """
        关闭Channel 1
        :return:返回关闭Channel的结果
        """
        self.toTxt("关闭当前Channel 1")
        self.gbs_refresh()

        self.openChannel_1_window()
        # 定位到 <video> 元素
        video_element = self.driver.find_element(By.TAG_NAME, 'video')
        # 使用 JavaScript 获取当前播放时间
        current_time = self.driver.execute_script('return arguments[0].currentTime;', video_element)
        # 打印当前播放时间
        self.toTxt(f'当前播放时间：{current_time} 秒')
        if str(current_time) == "0":
            self.toTxt("关闭Channel 1成功")
            return True
        else:
            self.toTxt("关闭Channel 1失败")
            return False

    def VideoChannel_2_Close(self):
        """
        关闭Channel 2
        :return:返回关闭Channel的结果
        """
        self.toTxt("关闭当前Channel 2")
        self.gbs_refresh()

        self.openChannel_2_window()
        # 定位到 <video> 元素
        video_element = self.driver.find_element(By.TAG_NAME, 'video')
        # 使用 JavaScript 获取当前播放时间
        current_time = self.driver.execute_script('return arguments[0].currentTime;', video_element)
        # 打印当前播放时间
        self.toTxt(f'当前播放时间：{current_time} 秒')
        if str(current_time) == "0":
            self.toTxt("关闭Channel 2成功")
            return True
        else:
            self.toTxt("关闭Channel 2失败")
            return False

    def openNewUrlBrowser(self, enter_url, index):
        self.toTxt("打开新的浏览器打开url：{}".format(enter_url))
        self.driver.execute_script("window.open('about:blank', 'new_tab')")
        # 切换到新标签页
        self.driver.switch_to.window(self.driver.window_handles[index])
        # 访问网页
        self.driver.get(enter_url)

    def checkDeviceStatusGBS(self):
        """
        检测GBS设备的在线状态
        :return:返回检查结果
        """
        self.toTxt("检查设备GBS是否在线")
        try:
            self.driver.find_element(By.XPATH, "//td[text()='ON']")
            self.toTxt("设备在线")
            return True
        except NoSuchElementException:
            self.toTxt("设备不在线")
            return False

    def OpenChannel1AndPlayVideo(self, ipcid):
        self.toTxt("打开Channel 1 并开始拉流，整合测试")
        self.openUrl("http://mc.seewo.com/cloud")
        waitTimeBack(1)
        self.loginGBS()
        if self.enterDeviceGBSVideoStreamControlPage(ipcid):
            self.openChannel_1_window()
            if self.Channel_1_Play():
                self.toTxt("Channel 1 播放成功！")
            else:
                self.toTxt("Channel 1 播放失败")
        else:
            self.toTxt("Channel 1 设备离线")

    # ------------------------ SXW0301Project Control End ------------------------

    def scrollToSpecificElement(self, find_type, element_content):
        """
        通过指定定位方式滚动到指定元素的位置
        :param find_type: 定位方式
        :param element_content: 定位内容
        :return:
        """
        self.toTxt("开始滚动查找元素……")
        target_element = ""
        # 定位到目标元素
        if find_type == "id":
            target_element = self.driver.find_element(By.ID, element_content)
        elif find_type == "xpath":
            target_element = self.driver.find_element(By.XPATH, element_content)
        elif find_type == "class":
            target_element = self.driver.find_element(By.CLASS_NAME, element_content)
        elif find_type == "text":
            target_element = self.driver.find_element(By.LINK_TEXT, element_content)
        # 使用JavaScript滚动到目标元素
        self.driver.execute_script("arguments[0].scrollIntoView();", target_element)
        self.toTxt("找到元素，进行下一步操作！")

    def killEdge(self):
        """
        杀掉Edge浏览器，并清空当前WebDriver
        :return:
        """
        self.toTxt("关闭浏览器进程")
        self.driver.quit()
        gc.collect()
        waitTimeBack(1)
        self.toTxt("关闭完成！\n\n\n")

    def toTxt(self, result):
        """
        运行时Log存储
        :param result:存储的log内容
        :return:
        """
        try:
            print(result)
            with open("{}/[{}]{}seleniumTestLog.log".format(os.path.dirname(os.path.realpath(__file__)),
                                                            self.log_saveName, self.log_name), "a+") as f:
                cur_time = time.strftime("%Y%m%d_%H%M%S")
                result = result.encode("utf-8", errors="ignore").decode("utf-8", errors="ignore")
                f.write("[{} - Device{}] - ".format(cur_time, self.log_saveName) + result + "\n")
        except (AttributeError, TypeError) as ex:
            print(
                "【Device{}】 - ".format(
                    self.log_saveName) + "【Error need check, maybe not important】 : \r\n{}\r\n".format(
                    str(ex)))
            f.write(
                "【Device{}】 - ".format(
                    self.log_saveName) + "【Error need check, maybe not important】 : \r\n{}\r\n".format(
                    str(ex)))


class multi_thread(threading.Thread):
    def __init__(self, *params, **known):
        super(multi_thread, self).__init__(*params, **known)
        parent_thread = threading.current_thread()
        self.is_killed = False
        self.child_threads = []
        if hasattr(parent_thread, 'child_threads'):
            parent_thread.child_threads.append(self)

    def _raise_exc(self, exc_obj):
        if not self.is_alive():
            return

        res = ctypes.pythonapi.PyThreadState_SetAsyncExc(
            ctypes.c_long(self.ident), ctypes.py_object(exc_obj))
        if res == 0:
            raise RuntimeError("Not existent thread id.")
        elif res > 1:
            ctypes.pythonapi.PyThreadState_SetAsyncExc(self.ident, None)
            raise SystemError("PyThreadState_SetAsyncExc failed.")

    def kill(self):
        if hasattr(self, 'child_threads'):
            for child_thread in self.child_threads:
                if child_thread.is_alive():
                    child_thread.kill()
        self._raise_exc(SystemExit)
        self.is_killed = True


"""
    Supply Comport Api
    继电器控制API
"""


class Supply_Comport_Api:

    def __init__(self, com_id, baudrate):
        self.com_id = com_id
        self.baud_rate = baudrate
        self.s_obj = serial.Serial(self.com_id, baudrate=self.baud_rate)

    # 根据每个开关对应的设备串口ID去打开该设备继电器电源打开
    def openOUT1(self, deviceid):
        print("Supply Control: {}".format(deviceid))
        """
        :param deviceid:设备的端口号
        :return:
        """
        global lock
        with lock:
            command = ""
            if deviceid == "COM1":
                command = RELAY_CONTROL_COMPORT_1_OPEN
            elif deviceid == "COM2":
                command = RELAY_CONTROL_COMPORT_2_OPEN
            elif deviceid == "COM3":
                command = RELAY_CONTROL_COMPORT_3_OPEN
            elif deviceid == "COM4":
                command = RELAY_CONTROL_COMPORT_4_OPEN
            elif deviceid == "COM5":
                command = RELAY_CONTROL_COMPORT_5_OPEN
            elif deviceid == "COM6":
                command = RELAY_CONTROL_COMPORT_6_OPEN
            elif deviceid == "COM7":
                command = RELAY_CONTROL_COMPORT_7_OPEN
            elif deviceid == "COM8":
                command = RELAY_CONTROL_COMPORT_8_OPEN
            elif deviceid == "COM9":
                command = RELAY_CONTROL_COMPORT_9_OPEN
            elif deviceid == "COM10":
                command = RELAY_CONTROL_COMPORT_10_OPEN
            self.s_obj.write(command)
            sleep(0.1)

    # 根据每个开关对应的设备串口ID去打开该设备继电器电源关闭
    def closeOUT1(self, deviceid):
        print("Supply Control: {}".format(deviceid))
        global lock
        with lock:
            command = ""
            if deviceid == "COM1":
                command = RELAY_CONTROL_COMPORT_1_CLOSE
            elif deviceid == "COM2":
                command = RELAY_CONTROL_COMPORT_2_CLOSE
            elif deviceid == "COM3":
                command = RELAY_CONTROL_COMPORT_3_CLOSE
            elif deviceid == "COM4":
                command = RELAY_CONTROL_COMPORT_4_CLOSE
            elif deviceid == "COM5":
                command = RELAY_CONTROL_COMPORT_5_CLOSE
            elif deviceid == "COM6":
                command = RELAY_CONTROL_COMPORT_6_CLOSE
            elif deviceid == "COM7":
                command = RELAY_CONTROL_COMPORT_7_CLOSE
            elif deviceid == "COM8":
                command = RELAY_CONTROL_COMPORT_8_CLOSE
            elif deviceid == "COM9":
                command = RELAY_CONTROL_COMPORT_9_CLOSE
            elif deviceid == "COM10":
                command = RELAY_CONTROL_COMPORT_10_CLOSE
            self.s_obj.write(command)
            sleep(0.1)

    # 短按Power键
    def shortPressPowerKey(self):
        print("短按Power键")
        self.openOUT1("COM1")
        self.closeOUT1("COM1")
        waitTimeBack(1)

    # 长按Power键
    def longPressPowerKey(self, long_click_time):
        print("长按Power键 {} 秒".format(long_click_time))
        self.openOUT1("COM1")
        waitTimeBack(long_click_time)
        self.closeOUT1("COM1")
        waitTimeBack(1)


"""
    Android Device Comport Api
    安卓设备对应串口API
"""


class SingleDeviceControlApi:

    def __init__(self, com_id, baud_rate, log_name):
        """
        单一串口控制初始化函数
        :param com_id:串口的id
        :param baud_rate: 串口波特率
        """
        self.com_id = com_id
        self.baud_rate = baud_rate
        self.port_obj = serial.Serial(self.com_id, self.baud_rate)
        self.log_name = log_name

    # 读取串口回传信息
    def readComport(self, data_size):
        """
        调用该函数会对当前串口信息进行读取并返回
        :return: 返回当前串口上报信息
        """
        # 每次刷入前都把缓存接收掉
        result = self.port_obj.read(data_size).decode("gbk", errors="ignore")
        return result

    # 实时读取串口回传信息
    def timing_reading(self):
        result_back = ""
        while True:
            data_size = self.port_obj.inWaiting()
            if data_size != 0:
                content = self.readComport(data_size)
                # 保证返回数据文本进行判断格式是原格式
                result_back += content
                print(content)
                # 保证写入log的数据是正常格式数据
                self.toTxt(content)
            else:
                return result_back
            sleep(0.1)

    # 写入串口指令
    def writeComport(self, command):
        """
        写入串口命令
        :param command:串口命令
        :return:无
        """
        self.port_obj.write(command.encode("UTF-8"))
        self.port_obj.flush()

    # 将当前串口数据写入Log
    def toTxt(self, result):
        """
        创建txt格式文本用于实时写入数据内容
        :param result:传入需要写入到文本的内容
        :return:无
        """
        cur_time = time.strftime("%Y%m%d_%H%M%S")
        global lock
        with lock:
            try:
                print(result)
                with open("./[{}]{}.log".format(self.port_obj.portstr, self.log_name), "a+", encoding="utf-8") as f:
                    f.write("[{}".format(cur_time) + "-{}]:".format(
                        self.com_id) + result + "\n")
            except (AttributeError, TypeError) as ex:
                print(
                    "[{}-".format(self.com_id) + "Error need check, maybe not important]: \r\n{}\r\n".format(
                        str(ex)))
                f.write(
                    "[{}-".format(self.com_id) + "Error need check, maybe not important]:\r\n{}\r\n".format(
                        str(ex)))


# 串口资源锁，放这里不动
lock = threading.RLock()


# 获取设备序列号 - adb
def getDeviceSerialNo():
    devices_stream = os.popen("adb devices")
    devices = devices_stream.read()
    devices_stream.close()
    serial_no = re.findall("(.*)\tdevice", devices)
    return serial_no


# 获取设备序列号 - fastboot - 需要都加上
def getDeviceFastbootSerialNo():
    devices_stream = os.popen("fastboot devices")
    devices = devices_stream.read()
    devices_stream.close()
    serial_no = re.findall("(.*)\tfastboot", devices)
    return serial_no


# write in device and its specific fastboot sn with adb sn into a file for reading - return file path
def initAllDeviceAndFastbootSN():
    adbSN_list = getDeviceSerialNo()
    print("当前有{}设备".format(adbSN_list))
    adbSN_fastbootSNDict = {}
    for device_sn_temp in adbSN_list:
        os.popen("adb -s {} reboot bootloader".format(device_sn_temp))
        print("设备进入bootloader中……")
        sleep(10)
        temp_fastbootSN = getDeviceFastbootSerialNo()[0]
        adbSN_fastbootSNDict[device_sn_temp] = temp_fastbootSN
        print(adbSN_fastbootSNDict[device_sn_temp])
        os.popen("fastboot -s {} reboot".format(temp_fastbootSN))
        print("设备进入adb中……")
        sleep(30)
    adb_fastboot_sn_fpath = "./adb_fastboot_sn_dict.json"
    with open(adb_fastboot_sn_fpath, "w") as f:
        print("正在写入不同设备对应的SN号序列号……")
        json.dump(adbSN_fastbootSNDict, f)
        f.close()
    return adb_fastboot_sn_fpath


# read from fastboot sn with adb sn file for test - return dict
def readAllDeviceAndFastbootSN(adb_fastboot_sn_fpath):
    with open(adb_fastboot_sn_fpath, "r") as f:
        serialDict = json.load(f)
        f.close()
    return serialDict


def getAllPorts():
    ports_device = []
    ports_supplyControl = []
    try:
        for port in list(comports()):
            # 继电器端口筛选
            if "Silicon Labs CP210x USB to UART Bridge" in str(port):
                current_port = re.findall("\((.*)\)", str(port))[0]
                ports_supplyControl.append(current_port)
            # 设备端口筛选
            if "USB Serial Port" in str(port):
                current_port = re.findall("\((.*)\)", str(port))[0]
                ports_device.append(current_port)
    except Exception:
        print("Serial port is lost, please check！")
    return ports_supplyControl, ports_device


"""
    结果生成区域：
    步骤1：生成excel空文件
    步骤2：通过自定义写入数据实现实时写入
"""


def standard_test_DataGenerate(form=""):
    df = pd.DataFrame({"测试次数": [""], "测试时间": [""], "结果": [""]})
    df.to_excel(form, engine="openpyxl")
    # 释放df资源实例
    df = ""


def write_into_excel(form="", sheet_name="Sheet1", row=1, column=12, testCount="1", testTime="", testResult="",
                     serialno=""):
    """
    通过openpyxl模块将每一行case的测试结果写入对应每一行的结果列中
    :param form:待写入case Excel文件路径
    :param sheet_name:待写入case Excel文件指定sheet表名
    :param row:待写入case测试结果所在行
    :param column:待写入case测试结果所在列
    :param value:待写入测试结果
    :return:None
    """
    print("将测试结果写入excel表格对应Case的行 - 测试结果处：【{}】".format(testResult))
    wb = openpyxl.load_workbook(form)
    ws = wb[sheet_name]
    grid_value_testCount = ws.cell(row + 1, column).value
    grid_value_testTime = ws.cell(row + 1, column + 1).value
    grid_value_testResult = ws.cell(row + 1, column + 2).value
    print("{} - {} - {}".format(testCount, testTime, testResult))
    if grid_value_testCount is None and grid_value_testTime is None and grid_value_testResult is None:
        ws.cell(row + 1, column).value = testCount
        ws.cell(row + 1, column + 1).value = testTime
        ws.cell(row + 1, column + 2).value = testResult
    wb.save(form)
    print("Done")


def is_valid_image(path):
    '''
    检查文件是否损坏
    '''
    try:
        bValid = True
        fileObj = open(path, 'rb')  # 以二进制形式打开
        buf = fileObj.read()
        if not buf.startswith(b'\xff\xd8'):  # 是否以\xff\xd8开头
            bValid = False
        elif buf[6:10] in (b'JFIF', b'Exif'):  # “JFIF”的ASCII码
            if not buf.rstrip(b'\0\r\n').endswith(b'\xff\xd9'):  # 是否以\xff\xd9结尾
                bValid = False
        else:
            try:
                Image.open(fileObj).verify()
            except Exception as e:
                bValid = False
                print(e)
    except Exception as e:
        return False
    return bValid


def waitTimeBack(timeSleepSecond):
    for i in range(1, timeSleepSecond + 1):
        print("已等待{}秒，还需等待{}秒".format(i, timeSleepSecond + 1 - i))
        sleep(1)
    print("等待结束！开始执行下一步操作~")
